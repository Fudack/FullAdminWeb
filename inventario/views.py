from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.db.models import Q
from .forms import ProductForm, ExcelUploadForm
from .models import Productos
from openpyxl.styles import NamedStyle, Font
from openpyxl import Workbook
import openpyxl

@login_required(login_url='login')
def export_data(response, data, data_titles, name):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename={}.xlsx'.format(name)
    
    wb = Workbook()
    ws = wb.active

    header_style = NamedStyle(name="header_style")
    header_style.font = Font(bold=True)
    wb.add_named_style(header_style)

    headers = [field for field in data_titles]

    for col_num, header in enumerate(headers[2:], start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.style = "header_style"

    data_style = NamedStyle(name="data_style")
    data_style.alignment.horizontal = "left"
    wb.add_named_style(data_style)

    for row_num, row in enumerate(data, start=2):
        print(row), print(row_num)
        for col_num, value in enumerate(row, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.style = "data_style"

    wb.save(response)
    return response

@login_required(login_url='login')
def product_list(request):
    excel_form = ExcelUploadForm()
    form = ProductForm()
    productos = Productos.objects.all()
    unique_categories = productos.values_list('categoria', flat=True).distinct()
    unique_ubicacion = productos.values_list('ubicacion', flat=True).distinct()
    data_titles = [field.name for field in Productos._meta.get_fields()]

    if request.GET.get('exportar'):
        data = [list(row.values()) for row in productos.values()]
        name = 'inventario'
        return export_data(request, data, data_titles, name)

    elif request.GET.get('formato'):
        data = []
        name = 'formato'
        return export_data(request, data, data_titles, name)

    elif request.GET.get('q'):
        query = request.GET.get('q')
        productos = Productos.objects.filter(
            Q(id__icontains=query) |
            Q(nombre__icontains=query) |
            Q(descripcion__icontains=query) |
            Q(precio_compra__icontains=query) |
            Q(precio_venta__icontains=query) |
            Q(stock__icontains=query) |
            Q(categoria__icontains=query) |
            Q(ubicacion__icontains=query)
        )

    elif  request.POST.get('importar'):
        excel_form = ExcelUploadForm(request.POST, request.FILES)
        print("hola")
        if excel_form.is_valid():
            excel_file = request.FILES['archivo']
            wb = openpyxl.load_workbook(excel_file)
            worksheet = wb.active
            for row in worksheet.iter_rows():
                nombre = row[0].value
                descripcion = row[1].value
                precio_compra = row[2].value
                precio_venta = row[3].value
                stock = row[4].value
                categoria = row[5].value
                ubicacion = row[6].value
                Productos.objects.create(
                    nombre=nombre,
                    descripcion=descripcion,
                    precio_compra=precio_compra,
                    precio_venta=precio_venta,
                    stock=stock,
                    categoria=categoria,
                    ubicacion=ubicacion
                )
            return redirect('product_list')

    elif request.GET.get('c'):
        filter_categories = request.GET.get('c')
        productos = Productos.objects.filter(
            Q(categoria__icontains=filter_categories)
        )

    # Rest of the filtering code remains the same

    context = {
        'productos': productos,
        'form': form,
        'unique_categories': unique_categories,
        'excel_form': excel_form
    }
    return render(request, 'inventario.html', context)

# Rest of the views remain unchanged

@login_required(login_url='login')
def contabilidad(request):
    return render(request, 'contabilidad.html')

@login_required(login_url='login')
def recursosHumanos(request):
    return render(request, 'recursosHumanos.html')

@login_required(login_url='login')
def facturacion(request):
    return render(request, 'facturacion.html')
